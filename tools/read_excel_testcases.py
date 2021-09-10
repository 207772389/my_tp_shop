# -*- coding: utf-8 -*-
"""
@Time ： 2021/9/9 10:08
@Auth ： cainiao
@File ：读取excel里的测试用例
@IDE ：PyCharm
@Motto：ABC(Always Be Coding)

"""
import openpyxl


def read_excel_cases():
    # 读取excel
    workbook = openpyxl.load_workbook('../data/test_cases.xlsx')
    # 指定你要读哪个sheel表的数据
    sheet = workbook['Sheet1']
    # 获取最大行数
    rows = sheet.max_row
    #方法一：
    list_data = []
    for line in range(2, rows + 1):
        # 通过dict转为字典格式的
        #   这个弊端就是：列数是写死的，如果excel修改了，这里的对应就会错乱。
        dict_case = dict(id=sheet.cell(row=line, column=1).value,
                         url=sheet.cell(row=line, column=2).value,
                         data=sheet.cell(row=line, column=3).value,
                         rep=sheet.cell(row=line, column=4).value),
        # 把每条字典放入列表
        list_data.append(dict_case)

    print(list_data)
    # 一定记得返回列表
    # return list_data

    # 方法二：提倡这个方法
    list_data2 = []
    for row in range(2, rows + 1):
        dict1={}
        for col in range(1,sheet.max_column+1):
            dict1[sheet.cell(1,col).value] = sheet.cell(row,col).value
        list_data2.append(dict1)

    print(list_data2)
    # 一定记得返回列表
    return list_data2


"""应用示例
#遍历最后的包含了字典的列表
    for data in list_data:
        id = data.get('id')
        url = data.get('url')
        #注意：这里记得转换格式，从excel文档里读出来的内容格式为数字或者字符串，必须转换下格式。
        data = eval(data.get('data'))
        rep = data.get('rep')

        #发送请求
        response = my_request.send(url,data)"""

if __name__ == '__main__':
    read_excel_cases()
